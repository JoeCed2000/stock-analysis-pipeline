"""Excel report generator — produces .xlsx with financial data, valuation, and risks."""
import os
import logging
from typing import Dict, Any, List
from datetime import datetime

logger = logging.getLogger(__name__)

_OPENPYXL_AVAILABLE = False
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    logger.warning("openpyxl not installed — Excel generation disabled")


if _OPENPYXL_AVAILABLE:
    # ── Style constants ──
    HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="16213e")
    SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="0f3460")
    NORMAL_FONT = Font(name="Calibri", size=11, color="1a1a2e")
    METRIC_FONT = Font(name="Calibri", size=11, bold=True, color="1a1a2e")
    GREEN_FONT = Font(name="Calibri", size=11, bold=True, color="1b7a1b")
    RED_FONT = Font(name="Calibri", size=11, bold=True, color="c0392b")
    AMBER_FONT = Font(name="Calibri", size=11, bold=True, color="e67e22")
    THIN_BORDER = Border(bottom=Side(style="thin", color="cccccc"))
    HEADER_BORDER = Border(bottom=Side(style="medium", color="0f3460"))


def _apply_header_style(ws, row: int, cols: int):
    """Apply header styling to a row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_section(ws, row: int, title: str, cols: int = 2) -> int:
    """Write a section title spanning columns. Returns next row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    return row + 1


def _write_row(ws, row: int, label: str, value, fmt: str = None, bold_value: bool = False) -> int:
    """Write a label-value row. Returns next row."""
    ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
    val_cell = ws.cell(row=row, column=2, value=value)
    val_cell.font = METRIC_FONT if bold_value else NORMAL_FONT
    if fmt:
        val_cell.number_format = fmt
    for col in range(1, 3):
        ws.cell(row=row, column=col).border = THIN_BORDER
    return row + 1


def _fmt_billions(val) -> str:
    """Format large number as billions string."""
    if val is None:
        return "N/A"
    if abs(val) >= 1e12:
        return f"${val/1e12:.2f}T"
    if abs(val) >= 1e9:
        return f"${val/1e9:.2f}B"
    if abs(val) >= 1e6:
        return f"${val/1e6:.1f}M"
    return f"${val:,.0f}"


def generate_excel(
    output_path: str,
    ticker: str,
    company_name: str,
    data: Dict[str, Any],
    risks: List[Dict[str, Any]] = None,
) -> str:
    """Generate a formatted Excel report. Returns the file path."""
    if not _OPENPYXL_AVAILABLE:
        logger.warning("openpyxl unavailable — skipping Excel generation")
        return ""

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1a1a2e"

    # Title
    ws.merge_cells("A1:C1")
    title_cell = ws.cell(row=1, column=1, value=f"{company_name} ({ticker}) — Financial Summary")
    title_cell.font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(name="Calibri", size=10, color="888888")

    row = 4
    _apply_header_style(ws, row, 3)
    ws.cell(row=row, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=row, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=row, column=3, value="Details").font = HEADER_FONT
    row += 1

    fin = data.get("financials", {})

    # Price & Market
    _write_section(ws, row, "💰 Price & Market", 3)
    row += 1
    row = _write_row(ws, row, "Current Price", data.get("price"), "#,##0.00")
    row = _write_row(ws, row, "Previous Close", data.get("prev_close"), "#,##0.00")
    row = _write_row(ws, row, "Market Cap", _fmt_billions(data.get("market_cap")))
    row = _write_row(ws, row, "Sector", data.get("sector"))
    row = _write_row(ws, row, "Industry", data.get("industry"))
    row = _write_row(ws, row, "Currency", data.get("currency"))
    row += 1

    # Revenue
    _write_section(ws, row, "📈 Revenue", 3)
    row += 1
    row = _write_row(ws, row, "Revenue (Annual)", _fmt_billions(fin.get("revenue_annual")))
    rev_growth = fin.get("revenue_annual_growth")
    row = _write_row(ws, row, "Revenue Growth (Annual)", f"{rev_growth*100:.1f}%" if rev_growth else "N/A")
    rev_q_growth = fin.get("revenue_yoy_growth")
    row = _write_row(ws, row, "Revenue Growth (YoY Q)", f"{rev_q_growth*100:.1f}%" if rev_q_growth else "N/A")
    row = _write_row(ws, row, "Revenue (Quarterly)", _fmt_billions(fin.get("revenue_quarterly")))
    row += 1

    # Profitability
    _write_section(ws, row, "💵 Profitability", 3)
    row += 1
    gm = fin.get("gross_margin")
    row = _write_row(ws, row, "Gross Margin", f"{gm*100:.1f}%" if gm else "N/A")
    om = fin.get("operating_margin")
    row = _write_row(ws, row, "Operating Margin", f"{om*100:.1f}%" if om else "N/A")
    row = _write_row(ws, row, "Net Income", _fmt_billions(fin.get("net_income")))
    row = _write_row(ws, row, "Free Cash Flow", _fmt_billions(fin.get("free_cash_flow")))
    row = _write_row(ws, row, "Net Debt", _fmt_billions(fin.get("net_debt")))
    row += 1

    # Valuation
    _write_section(ws, row, "📊 Valuation", 3)
    row += 1
    pe = data.get("pe_current")
    fpe = data.get("pe_forward")
    row = _write_row(ws, row, "P/E (Trailing)", f"{pe:.1f}" if pe else "N/A")
    row = _write_row(ws, row, "P/E (Forward)", f"{fpe:.1f}" if fpe else "N/A")
    peg = data.get("peg_ratio")
    row = _write_row(ws, row, "PEG Ratio", f"{peg:.2f}" if peg else "N/A")
    row = _write_row(ws, row, "Beta", f"{data.get('beta'):.2f}" if data.get("beta") else "N/A")
    row = _write_row(ws, row, "52-Week High", data.get("52w_high"), "#,##0.00")
    row = _write_row(ws, row, "52-Week Low", data.get("52w_low"), "#,##0.00")

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18

    # ── Sheet 2: Valuation Details ──
    ws2 = wb.create_sheet("Valuation")
    ws2.sheet_properties.tabColor = "0f3460"

    ws2.merge_cells("A1:B1")
    ws2.cell(row=1, column=1, value=f"{ticker} — Valuation Analysis").font = TITLE_FONT
    ws2.row_dimensions[1].height = 30

    row2 = 3
    _write_section(ws2, row2, "Valuation Ratios", 2)
    row2 += 1

    pe_val = fpe if fpe else pe
    row2 = _write_row(ws2, row2, "P/E Ratio", f"{pe_val:.1f}" if pe_val else "N/A")
    row2 = _write_row(ws2, row2, "PEG Ratio", f"{peg:.2f}" if peg else "N/A")
    eg = data.get("expected_growth")
    row2 = _write_row(ws2, row2, "Expected Growth", f"{eg*100:.1f}%" if eg else "N/A")
    row2 += 1

    # Margin of Safety
    _write_section(ws2, row2, "Margin of Safety", 2)
    row2 += 1
    if pe_val:
        if pe_val < 15:
            mos = "✅ COMFORTABLE (PE < 15)"
            color = GREEN_FONT
        elif pe_val < 25:
            mos = "⚠️ MODERATE (PE 15-25)"
            color = AMBER_FONT
        elif pe_val < 40:
            mos = "⚠️ WEAK (PE 25-40)"
            color = AMBER_FONT
        else:
            mos = "🔴 NONE (PE > 40)"
            color = RED_FONT
        row2 = _write_row(ws2, row2, "Assessment", mos)
        ws2.cell(row=row2 - 1, column=2).font = color

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 30

    # ── Sheet 3: Risk Assessment ──
    if risks:
        ws3 = wb.create_sheet("Risks")
        ws3.sheet_properties.tabColor = "c0392b"

        ws3.merge_cells("A1:D1")
        ws3.cell(row=1, column=1, value=f"{ticker} — Risk Assessment").font = TITLE_FONT
        ws3.row_dimensions[1].height = 30

        row3 = 3
        _apply_header_style(ws3, row3, 4)
        for col_idx, header in enumerate(["Category", "Description", "Severity", "Source"], 1):
            ws3.cell(row=row3, column=col_idx, value=header)
        row3 += 1

        severity_colors = {
            "high": RED_FONT,
            "medium": AMBER_FONT,
            "low": GREEN_FONT,
        }

        for risk in risks:
            if isinstance(risk, dict):
                cat, desc, sev, src = risk.get("category",""), risk.get("description",""), risk.get("severity",""), risk.get("source","")
            else:
                cat, desc, sev, src = risk.category, risk.description, risk.severity, risk.source
            ws3.cell(row=row3, column=1, value=cat).font = METRIC_FONT
            ws3.cell(row=row3, column=2, value=desc).font = NORMAL_FONT
            sev_cell = ws3.cell(row=row3, column=3, value=sev.upper() if sev else "")
            sev_cell.font = severity_colors.get(sev.lower(), NORMAL_FONT)
            ws3.cell(row=row3, column=4, value=src).font = Font(name="Calibri", size=10, color="888888")
            for col in range(1, 5):
                ws3.cell(row=row3, column=col).border = THIN_BORDER
            row3 += 1

        ws3.column_dimensions["A"].width = 18
        ws3.column_dimensions["B"].width = 55
        ws3.column_dimensions["C"].width = 12
        ws3.column_dimensions["D"].width = 35

    # Save
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    logger.info(f"Excel saved: {output_path} ({', '.join(wb.sheetnames)} sheets)")
    return output_path
